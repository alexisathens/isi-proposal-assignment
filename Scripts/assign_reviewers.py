#!/usr/bin/env python3
"""
ips_assign_reviewers.py

Purpose:
  Assign IPS proposals to reviewers ensuring:
  1. ALL reviewers from an association review ALL papers from that association
  2. Every proposal gets exactly 2 reviewers
  3. Workload is distributed as equitably as possible (no fixed limit)

Example:
  python ips_assign_reviewers.py \
    --proposals proposals_with_embeddings.json \
    --members members_with_embeddings.json \
    --out assignment_results.xlsx
"""
from __future__ import annotations

import os
import json
import argparse
import pandas as pd
import logging
from typing import Dict, Any, List, Tuple
from collections import defaultdict, Counter

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# initialize logging (small and local)
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

# New helper: ensure similarities are in the shape {pid: {rid: score}}
def _ensure_all_sims_mapping(all_sims):
	"""
	Normalize common shapes into {proposal_id: {member_id: score, ...}}.
	Accepts dicts, lists of dicts, row-wise lists, etc.
	"""
	if all_sims is None:
		return {}
	# Already in desired shape
	if isinstance(all_sims, dict):
		out = {}
		for pid, val in all_sims.items():
			if isinstance(val, dict):
				out[str(pid)] = {str(k): float(v) for k, v in val.items()}
			elif isinstance(val, list):
				# try to coerce list entries into dict
				nested = {}
				for it in val:
					if isinstance(it, dict):
						mid = it.get("Member ID") or it.get("member_id") or it.get("Reviewer ID") or it.get("rid") or it.get("ID")
						score = it.get("score") or it.get("similarity") or it.get("value")
						if mid is not None:
							nested[str(mid)] = float(score or 0)
				out[str(pid)] = nested
			else:
				out[str(pid)] = {}
		return out

	# List of entries -> coerce row-wise or grouped shapes
	if isinstance(all_sims, list):
		out = {}
		for entry in all_sims:
			if not isinstance(entry, dict):
				continue
			# grouped: {"Proposal ID": pid, "scores": {rid:score, ...}}
			if "Proposal ID" in entry and "scores" in entry and isinstance(entry["scores"], dict):
				pid = str(entry["Proposal ID"])
				out[pid] = {str(k): float(v) for k, v in entry["scores"].items()}
				continue
			# row-wise: {"Proposal ID": pid, "Member ID": mid, "score": s}
			pid = entry.get("Proposal ID") or entry.get("proposal_id") or entry.get("pid") or entry.get("ProposalID")
			mid = entry.get("Member ID") or entry.get("member_id") or entry.get("Reviewer ID") or entry.get("rid") or entry.get("ID")
			score = entry.get("score") or entry.get("similarity") or entry.get("value")
			if pid is not None and mid is not None:
				pid = str(pid); mid = str(mid)
				out.setdefault(pid, {})[mid] = float(score or 0)
				continue
			# nested list of scores: {"Proposal ID": pid, "scores": [ { "Member ID": mid, "score": s }, ... ]}
			if "Proposal ID" in entry and "scores" in entry and isinstance(entry["scores"], list):
				pid = str(entry["Proposal ID"])
				out.setdefault(pid, {})
				for srec in entry["scores"]:
					if isinstance(srec, dict):
						mid = srec.get("Member ID") or srec.get("member_id") or srec.get("ID")
						score = srec.get("score") or srec.get("similarity")
						if mid is not None:
							out[pid][str(mid)] = float(score or 0)
				continue
			logging.debug("Unrecognized similarity entry shape: %s", list(entry.keys()))
		return out

	logging.warning("Cannot normalize all_sims of type %s; returning empty mapping", type(all_sims))
	return {}


def calculate_all_similarities(proposals, members):
    """
    Compute pairwise cosine similarities between proposal embeddings and member embeddings.
    Returns: dict { proposal_id: { member_id: similarity, ... }, ... }
    """
    sims: Dict[str, Dict[str, float]] = {}

    # Build member embeddings map
    member_emb_map: Dict[str, np.ndarray] = {}
    for m in members:
        mid = str(m.get("ID", "")).strip()
        raw = m.get("embedding") or m.get("emb") or m.get("Embedding")
        if isinstance(raw, str):
            try:
                raw = json.loads(raw)
            except Exception:
                logging.warning("Member %s embedding is a string but JSON parsing failed", mid)
                raw = None
        if raw:
            try:
                member_emb_map[mid] = np.array(raw, dtype=float)
            except Exception:
                logging.warning("Failed to convert member %s embedding to numpy array", mid)

    if not member_emb_map:
        logging.warning("No member embeddings found; all similarities will be 0")

    # Compute similarities
    for p in proposals:
        pid = str(p.get("Proposal ID", "")).strip()
        rawp = p.get("embedding") or p.get("emb") or p.get("Embedding")
        if isinstance(rawp, str):
            try:
                rawp = json.loads(rawp)
            except Exception:
                logging.warning("Proposal %s embedding is a string but JSON parsing failed", pid)
                rawp = None
        if not rawp:
            logging.warning("No embedding for proposal %s -> similarities will be 0 for this proposal", pid)
            sims[pid] = {mid: 0.0 for mid in member_emb_map.keys()}
            continue
        try:
            pvec = np.array(rawp, dtype=float)
        except Exception:
            logging.warning("Failed to convert proposal %s embedding to numpy array", pid)
            sims[pid] = {mid: 0.0 for mid in member_emb_map.keys()}
            continue

        sims[pid] = {}
        for mid, mvec in member_emb_map.items():
            try:
                sims[pid][mid] = cos_sim(pvec, mvec)
            except Exception:
                sims[pid][mid] = 0.0

    return sims


def cos_sim(v1: np.ndarray, v2: np.ndarray) -> float:
    """Calculate cosine similarity between two vectors."""
    a = np.linalg.norm(v1)
    b = np.linalg.norm(v2)
    if a == 0 or b == 0:
        return 0.0
    return float(np.dot(v1, v2) / (a * b))


def get_member_meta(reviewer_id: str, members: List[Dict[str, Any]]) -> Tuple[str, str, str]:
    """Get member name, expertise, and association."""
    for m in members:
        if str(m.get("ID", "")) == reviewer_id:
            name = str(m.get("Members", ""))
            expertise = str(m.get("Expertise", ""))
            association = str(m.get("Association", m.get("Association", "")))
            return name, expertise, association
    return "", "", ""


def assign_reviewers_two_rounds(
    proposals: List[Dict[str, Any]], 
    members: List[Dict[str, Any]],
    sims=None
) -> Tuple[Dict[str, List[str]], Dict[str, Dict[str, float]]]:
    """
    Assign reviewers ensuring:
    1. ALL reviewers from an association review ALL papers from that association
    2. Every proposal gets exactly 2 reviewers
    3. Workload is distributed as equitably as possible
    
    Returns:
        - assignments: {proposal_id: [reviewer_id1, reviewer_id2]}
        - sim_scores: {proposal_id: {reviewer_id: similarity_score}}
    """
    
    # normalize sims into the mapping expected by the rest of the function
    sims_map = _ensure_all_sims_mapping(sims if sims is not None else globals().get("all_sims", None))

    # Calculate target workload dynamically
    total_proposals = len(proposals)
    total_members = len(members)
    total_reviews_needed = total_proposals * 2
    base_target = total_reviews_needed // total_members
    extra_reviews = total_reviews_needed % total_members
    
    print(f"\n=== Workload Calculation ===")
    print(f"Total proposals: {total_proposals}")
    print(f"Total members: {total_members}")
    print(f"Total reviews needed: {total_reviews_needed}")
    print(f"Base reviews per member: {base_target}")
    print(f"Members who need +1 extra: {extra_reviews}")
    print(f"Target distribution: {total_members - extra_reviews} members with {base_target} reviews, {extra_reviews} members with {base_target + 1} reviews")

    # Initialize structures
    assignments: Dict[str, List[str]] = {}
    reviewer_counts: Dict[str, int] = {str(m.get("ID", "")): 0 for m in members}
    sim_scores: Dict[str, Dict[str, float]] = defaultdict(dict)
    
    # Initialize all proposals in assignments
    for p in proposals:
        pid = str(p.get("Proposal ID", ""))
        assignments[pid] = []
    
    # Define specific associations (not "Other")
    SPECIFIC_ASSOCIATIONS = {"BS", "IAOS", "IASC", "IASE", "IASS", "IFC", "ISBIS", "ISI", "KOSTAT", "TIES", "Women", "Young"}
    
    # Create mappings
    member_by_assoc: Dict[str, List[str]] = defaultdict(list)
    members_with_other = []
    
    print("\nDEBUG - Processing members:")
    for m in members:
        mid = str(m.get("ID", ""))
        assoc = str(m.get("Association", m.get("Association", ""))).strip()
        name = str(m.get("Members", ""))
        print(f"  ID={mid}, Name={name}, Association='{assoc}'")
        member_by_assoc[assoc].append(mid)
        
        if assoc == "Other":
            members_with_other.append(mid)
    
    proposals_by_assoc: Dict[str, List[str]] = defaultdict(list)
    proposals_with_specific_assoc = []
    proposals_with_other = []
    
    print("\nDEBUG - Processing proposals:")
    for p in proposals:
        pid = str(p.get("Proposal ID", ""))
        title = str(p.get("Title", ""))[:50]
        assoc = str(p.get("Association", "")).strip()
        print(f"  ID={pid}, Title={title}, Association='{assoc}'")
        proposals_by_assoc[assoc].append(pid)
        
        if assoc in SPECIFIC_ASSOCIATIONS:
            proposals_with_specific_assoc.append(pid)
        elif assoc == "Other":
            proposals_with_other.append(pid)
    
    print(f"\n=== ROUND 1: Association-based Assignment ===")
    print(f"\nProposals WITH specific association: {len(proposals_with_specific_assoc)}")
    print(f"Proposals with 'Other' association: {len(proposals_with_other)}")
    print(f"Members from specific associations: {sum(len(member_by_assoc[a]) for a in SPECIFIC_ASSOCIATIONS if a in member_by_assoc)}")
    print(f"Members with 'Other' association: {len(members_with_other)}")
    
    # Round 1 - Part A: ALL members from an association review ALL proposals from that association
    print(f"\n--- Round 1A: Assigning proposals by specific association (ALL members review ALL papers) ---")
    
    for assoc in sorted(SPECIFIC_ASSOCIATIONS):
        if assoc not in proposals_by_assoc:
            continue
            
        pids = proposals_by_assoc[assoc]
        assoc_members = member_by_assoc.get(assoc, [])
        
        print(f"\n{assoc}: {len(pids)} proposal(s), {len(assoc_members)} member(s)")
        
        if not assoc_members:
            print(f"  ⚠ WARNING: No members found for association '{assoc}'")
            continue
        
        # Every member from this association reviews EVERY proposal from this association
        for pid in pids:
            for rid in assoc_members:
                assignments[pid].append(rid)
                reviewer_counts[rid] += 1
                
                sim = sims_map.get(pid, {}).get(rid, 0)
                sim_scores[pid][rid] = sim
                
                name, _, member_assoc = get_member_meta(rid, members)
                print(f"  ✓ Proposal {pid} → {name} ({member_assoc}) [Round 1A: association match, sim={sim:.3f}]")
        
        # Report final counts for this association
        for rid in assoc_members:
            name, _, _ = get_member_meta(rid, members)
            count = reviewer_counts[rid]
            print(f"  ℹ {name} has {count} proposals")
    
    round1a_assignments = sum(len(revs) for revs in assignments.values())
    print(f"\n✓ Round 1A complete: {round1a_assignments} assignments made")
    
    # Round 1 - Part B: Assign "Other" proposals to "Other" members using embeddings
    print(f"\n--- Round 1B: Assigning 'Other' proposals (first reviewer, using embeddings) ---")
    
    if proposals_with_other and members_with_other:
        available_proposals = set([p for p in proposals_with_other if len(assignments[p]) == 0])
        
        print(f"  'Other' proposals needing assignment: {len(available_proposals)}")
        
        while available_proposals:
            # Find best proposal-member pair using embeddings
            # Prioritize members with lower counts for fairness
            best_pair = None  # (pid, mid, sim, count)
            for pid in list(available_proposals):
                for mid in members_with_other:
                    sim = sims_map.get(pid, {}).get(mid, 0)
                    count = reviewer_counts[mid]
                    # Use a score that balances similarity and fairness
                    # Lower count is better (negative weight)
                    fairness_score = sim - (count * 0.1)  # Penalize high counts
                    
                    if best_pair is None or fairness_score > best_pair[3]:
                        best_pair = (pid, mid, sim, fairness_score)
            
            if best_pair is None:
                break
            
            pid_best, mid_best, sim_best, _ = best_pair
            assignments[pid_best].append(mid_best)
            reviewer_counts[mid_best] += 1
            sim_scores[pid_best][mid_best] = sim_best
            available_proposals.remove(pid_best)
            
            name, _, _ = get_member_meta(mid_best, members)
            print(f"  ✓ Proposal {pid_best} → {name} [Round 1B: sim={sim_best:.3f}] (reviewer now has {reviewer_counts[mid_best]} proposals)")
    elif proposals_with_other:
        print(f"  ⚠ WARNING: {len(proposals_with_other)} 'Other' proposals but no 'Other' members available")
    else:
        print(f"  No 'Other' proposals to assign")
    
    print(f"\n=== Round 1 Complete ===")
    total_r1 = sum(len(revs) for revs in assignments.values())
    print(f"Total assignments after Round 1: {total_r1}")
    
    # Show reviewer distribution after Round 1
    print("\nReviewer counts after Round 1:")
    for assoc in sorted(set(member_by_assoc.keys())):
        assoc_members = member_by_assoc[assoc]
        for mid in assoc_members:
            name, _, _ = get_member_meta(mid, members)
            count = reviewer_counts[mid]
            if count > 0:
                print(f"  {name} ({assoc}): {count} proposals")
    
    # Round 2: Ensure every proposal has exactly 2 reviewers
    print(f"\n=== ROUND 2: Second Reviewer Assignment (ensuring complete coverage) ===")
    
    need_second = [pid for pid, revs in assignments.items() if len(revs) < 2]
    all_member_ids = list(reviewer_counts.keys())
    
    print(f"\nProposals needing second reviewer: {len(need_second)}")
    
    # Sort proposals by how many reviewers they still need (prioritize those with 0)
    need_second.sort(key=lambda pid: len(assignments[pid]))
    
    # Iteratively assign second reviewers
    # Strategy: Always pick the reviewer with the lowest count who hasn't reviewed this proposal
    iteration = 0
    max_iterations = len(need_second) * len(all_member_ids)  # Prevent infinite loops
    
    while need_second and iteration < max_iterations:
        iteration += 1
        
        pid = need_second[0]  # Take first proposal that needs assignment
        
        # Find eligible members (not already assigned to this proposal)
        eligible = [(mid, reviewer_counts[mid], sims_map.get(pid, {}).get(mid, 0)) 
                   for mid in all_member_ids 
                   if mid not in assignments[pid]]
        
        if not eligible:
            print(f"  ⚠ ERROR: No eligible reviewers for proposal {pid}")
            need_second.pop(0)
            continue
        
        # Sort by: 1) current count (ascending), 2) similarity (descending)
        eligible.sort(key=lambda x: (x[1], -x[2]))
        
        mid_best, count, sim_best = eligible[0]
        
        assignments[pid].append(mid_best)
        reviewer_counts[mid_best] += 1
        sim_scores[pid][mid_best] = sim_best
        
        name, _, assoc = get_member_meta(mid_best, members)
        
        if iteration <= 20 or iteration % 10 == 0:  # Print first 20, then every 10th
            print(f"  ✓ Proposal {pid} → {name} ({assoc}) [sim={sim_best:.3f}] (reviewer now has {reviewer_counts[mid_best]} proposals)")
        
        # Check if this proposal is now complete
        if len(assignments[pid]) >= 2:
            need_second.remove(pid)
    
    if need_second:
        print(f"\n⚠ WARNING: {len(need_second)} proposals still need a second reviewer")
        print(f"  Proposal IDs: {need_second}")
    else:
        print(f"\n✓ All proposals have 2 reviewers!")
    
    print(f"\n=== Round 2 Complete ===")
    total_assignments = sum(len(revs) for revs in assignments.values())
    print(f"Total assignments: {total_assignments}")
    
    # Final workload analysis
    print(f"\n=== Final Workload Distribution ===")
    workloads = list(reviewer_counts.values())
    print(f"Min: {min(workloads)} | Max: {max(workloads)} | Avg: {sum(workloads)/len(workloads):.2f}")
    print(f"Target was: {base_target}-{base_target+1} reviews per member")
    
    # Show members who are significantly over/under target
    over_target = [(mid, reviewer_counts[mid]) for mid in all_member_ids 
                   if reviewer_counts[mid] > base_target + 1]
    under_target = [(mid, reviewer_counts[mid]) for mid in all_member_ids 
                    if reviewer_counts[mid] < base_target]
    
    if over_target:
        print(f"\nMembers above target ({base_target + 1}):")
        for mid, count in sorted(over_target, key=lambda x: -x[1])[:10]:
            name, _, assoc = get_member_meta(mid, members)
            print(f"  {name} ({assoc}): {count} reviews")
    
    if under_target:
        print(f"\nMembers below target ({base_target}):")
        for mid, count in sorted(under_target, key=lambda x: x[1])[:10]:
            name, _, assoc = get_member_meta(mid, members)
            print(f"  {name} ({assoc}): {count} reviews")
    
    return dict(assignments), dict(sim_scores)


def main() -> None:
    ap = argparse.ArgumentParser(description="Assign reviewers to IPS proposals ensuring complete coverage")
    ap.add_argument("--proposals", default="proposals_with_embeddings.json")
    ap.add_argument("--members", default="members_with_embeddings.json")
    ap.add_argument("--out", default="assignment_results.xlsx")
    args = ap.parse_args()

    # Load proposals
    proposals_list = []
    if os.path.exists(args.proposals):
        if args.proposals.lower().endswith(".json"):
            with open(args.proposals, "r", encoding="utf-8") as f:
                proposals_list = json.load(f)
        else:
            df = pd.read_csv(args.proposals, sep=';', encoding='latin1', dtype=str).fillna('')
            proposals_list = df.to_dict(orient='records')
        
    for r in proposals_list:
        if 'Association' in r and ('Association' not in r or not r.get('Association')):
            r['Association'] = r['Association']
        if not r.get('Association', '').strip():
            r['Association'] = 'Other'
        emb = r.get("embedding")
        if isinstance(emb, str) and emb.startswith('['):
            try:
                r["embedding"] = json.loads(emb)
            except Exception:
                pass
    else:
        logging.warning("Proposals file not found: %s", args.proposals)

    # Load members
    members_list = []
    if os.path.exists(args.members):
        if args.members.lower().endswith(".json"):
            with open(args.members, "r", encoding="utf-8") as f:
                members_list = json.load(f)
        else:
            df = pd.read_csv(args.members, sep=';', encoding='latin1', dtype=str).fillna('')
            members_list = df.to_dict(orient='records')
        
    for r in members_list:
        if 'Association' in r and ('Association' not in r or not r.get('Association')):
            r['Association'] = r['Association']
        if not r.get('Association', '').strip():
            r['Association'] = 'Other'
        emb = r.get("embedding")
        if isinstance(emb, str) and emb.startswith('['):
            try:
                r["embedding"] = json.loads(emb)
            except Exception:
                pass
    else:
        logging.warning("Members file not found: %s", args.members)

    print(f"Loaded {len(proposals_list)} proposals and {len(members_list)} members")

    # Compute similarities
    all_sims = calculate_all_similarities(proposals_list, members_list)
    norm_sims = _ensure_all_sims_mapping(all_sims)

    # Assign reviewers
    assignments, sim_scores = assign_reviewers_two_rounds(proposals_list, members_list, sims=norm_sims)

    # Write Excel
    print(f"\n=== Writing results to {args.out} ===")
    
    # Sheet 1: Assignments by Proposal
    proposals_data = []
    for pid, revs in assignments.items():
        p = next((x for x in proposals_list if str(x.get("Proposal ID")) == pid), {})
        
        description = ""
        if p.get("Description"):
            desc_dict = p.get("Description", {})
            brief = desc_dict.get("Brief Description", "")
            detailed = desc_dict.get("WSC Proposal Description", "")
            description = f"{brief}\n{detailed}".strip()
        
        proposal_assoc = p.get("Association", "")
        row = {
            "Proposal ID": pid,
            "Proposal Title": p.get("Title", ""),
            "Description": description,
            "Category": p.get("Category", ""),
            "Proposal Association": proposal_assoc
        }
        
        for i, rid in enumerate(revs, 1):
            name, expertise, assoc = get_member_meta(rid, members_list)
            row[f"Reviewer {i} ID"] = rid
            row[f"Reviewer {i} Name"] = name
            row[f"Reviewer {i} Association"] = assoc
            row[f"Reviewer {i} Expertise"] = expertise
            row[f"Similarity Score {i}"] = round(sim_scores.get(pid, {}).get(rid, 0), 3)
        
        proposals_data.append(row)

    df_proposals = pd.DataFrame(proposals_data)
    
    # Sheet 2: Assignments by Reviewer
    reviewer_counts: Dict[str, int] = defaultdict(int)
    for revs in assignments.values():
        for r in revs:
            reviewer_counts[r] += 1
    
    reviewer_data = []
    for mid in reviewer_counts.keys():
        name, expertise, assoc = get_member_meta(mid, members_list)
        assigned_proposals = []
        proposal_assocs = []
        similarities = []
        
        for pid, revs in assignments.items():
            if mid in revs:
                p = next((x for x in proposals_list if str(x.get("Proposal ID")) == pid), {})
                prop_assoc = p.get("Association", "")
                proposal_assocs.append(prop_assoc)
                assigned_proposals.append(pid)
                similarities.append(round(sim_scores.get(pid, {}).get(mid, 0), 3))
        
        reviewer_data.append({
            "Reviewer ID": mid,
            "Reviewer Name": name,
            "Reviewer Association": assoc if assoc else "No Association",
            "Expertise": expertise,
            "Number of Proposals": len(assigned_proposals),
            "Proposal IDs": ", ".join(assigned_proposals),
            "Proposal Associations": ", ".join(proposal_assocs),
            "Average Similarity": round(sum(similarities) / len(similarities), 3) if similarities else 0
        })
    
    df_reviewers = pd.DataFrame(reviewer_data)
    if not df_reviewers.empty and "Reviewer Association" in df_reviewers.columns:
        df_reviewers = df_reviewers.sort_values(["Reviewer Association", "Reviewer Name"])
    
    # Sheet 3: Detailed Assignments
    detailed_data = []
    for pid, revs in assignments.items():
        p = next((x for x in proposals_list if str(x.get("Proposal ID")) == pid), {})
        
        description = ""
        if p.get("Description"):
            desc_dict = p.get("Description", {})
            brief = desc_dict.get("Brief Description", "")
            detailed = desc_dict.get("WSC Proposal Description", "")
            description = f"{brief}\n{detailed}".strip()
        
        prop_assoc = p.get("Association", "")
        for rid in revs:
            name, expertise, assoc = get_member_meta(rid, members_list)
            detailed_data.append({
                "Proposal ID": pid,
                "Proposal Title": p.get("Title", ""),
                "Description": description,
                "Proposal Category": p.get("Category", ""),
                "Proposal Association": prop_assoc,
                "Reviewer ID": rid,
                "Reviewer Name": name,
                "Reviewer Association": assoc if assoc else "No Association",
                "Reviewer Expertise": expertise,
                "Similarity Score": round(sim_scores.get(pid, {}).get(rid, 0), 3)
            })

    df_detailed = pd.DataFrame(detailed_data)
    if not df_detailed.empty:
        df_detailed = df_detailed.sort_values(["Proposal Association", "Proposal ID", "Similarity Score"], ascending=[True, True, False])
    
    # Write to Excel
    with pd.ExcelWriter(args.out, engine='openpyxl') as writer:
        df_proposals.to_excel(writer, sheet_name='By Proposal', index=False)
        df_reviewers.to_excel(writer, sheet_name='By Reviewer', index=False)
        df_detailed.to_excel(writer, sheet_name='Detailed Assignments', index=False)
    
    # Format Excel
    wb = load_workbook(args.out)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        ws.freeze_panes = 'A2'
    
    wb.save(args.out)

    # Print detailed statistics
    print("\n=== Assignment Statistics ===")
    
    # Proposal coverage
    proposals_with_0 = sum(1 for revs in assignments.values() if len(revs) == 0)
    proposals_with_1 = sum(1 for revs in assignments.values() if len(revs) == 1)
    proposals_with_2 = sum(1 for revs in assignments.values() if len(revs) == 2)
    proposals_with_more = sum(1 for revs in assignments.values() if len(revs) > 2)
    
    print(f"\nProposal coverage:")
    print(f"  Proposals with 0 reviewers: {proposals_with_0}")
    print(f"  Proposals with 1 reviewer: {proposals_with_1}")
    print(f"  Proposals with 2 reviewers: {proposals_with_2}")
    if proposals_with_more > 0:
        print(f"  Proposals with >2 reviewers: {proposals_with_more}")
    
    # Reviewer workload
    total_assigned = sum(reviewer_counts.values())
    print(f"  Total assignments: {total_assigned}")
    
    if reviewer_counts:
        workloads = list(reviewer_counts.values())
        print(f"\nReviewer workload:")
        print(f"  Min proposals per reviewer: {min(workloads)}")
        print(f"  Max proposals per reviewer: {max(workloads)}")
        print(f"  Average proposals per reviewer: {sum(workloads)/len(workloads):.1f}")
        
        # Distribution
        dist = Counter(workloads)
        print(f"\n  Workload distribution:")
        for num_proposals in sorted(dist.keys()):
            num_reviewers = dist[num_proposals]
            print(f"    {num_reviewers} reviewers have {num_proposals} proposal(s)")
    
    # Average similarity scores
    all_similarities = []
    for pid, scores in sim_scores.items():
        all_similarities.extend(scores.values())
    
    if all_similarities:
        avg_sim = sum(all_similarities) / len(all_similarities)
        min_sim = min(all_similarities)
        max_sim = max(all_similarities)
        print(f"\nSimilarity scores:")
        print(f"  Average: {avg_sim:.3f}")
        print(f"  Min: {min_sim:.3f}")
        print(f"  Max: {max_sim:.3f}")
    
    print(f"\n✓ Assignments saved to {args.out}")


if __name__ == "__main__":
    main()